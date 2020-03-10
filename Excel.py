from openpyxl.styles import PatternFill
from openpyxl import load_workbook
import os
import re
import datetime

try:
    file = input('------------请输入统计表路径------------\n')
    folder = input('------------统计人员文件夹----------\n')
    red_fill = PatternFill("solid", fgColor="FFA500")
    green_fill = PatternFill("solid", fgColor="98FB98")
    # file = 'C:\\Users\Administrator\Desktop\\04英美文化分组.xlsx'
    # folder = 'C:\\Users\Administrator\Desktop\\英美比较\\'
    if folder[-1] != "\\":
        folder = folder + '\\'

    lists = os.listdir(folder)
except:
    print('输入路径有错误，请检查(不允许有其它字符)')
    input()
    exit()
alls = []
for name in lists:
    number = re.findall('\d+', name)
    for strs in number:
        if len(strs) == 11:
            number = strs
    alls.append([name, number])

print('所有交作业人员列表\n', alls)

copy_file = file.split('.')
copy_file = copy_file[0] + ' 输出结果 ' + str(datetime.date.today()) + '.' + copy_file[1]

data = load_workbook(file, data_only=True)
datas = data.active
scope = datas.rows
nones = 0
for row in scope:
    for val in row:
        if val.value == None:
            nones += 1
            break
        if len(str(val.value)) == 11:
            val.value = str(val.value)
            try:
                if_exist = False
                int(val.value)
                for i in range(len(alls)):
                    if str(alls[i][1]) == str(val.value):
                        if_exist = True
                        break
                if if_exist:
                    print(val.value + ':作业已交')
                    for vals in row:
                        vals.fill = green_fill
                else:
                    print(val.value + ':作业未交')
                    for vals in row:
                        vals.fill = red_fill
            except:
                print(val.value)
                pass
    if nones >= 300:
        break
print('-------正在写入文件-------')
print('-------该过程将持续30s左右-------')
data.save(copy_file)

print('-------写入成功，请查看目录-------')
input()
